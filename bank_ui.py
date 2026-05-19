import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
from datetime import datetime

# UPDATED: Using v2 files to prevent crashing with old data structures
EXCEL_FILE = "bank_data_v2.xlsx"
TXT_FILE = "bank_log_v2.txt"

class BankApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Banking System")
        self.root.configure(bg="#f4f5f7")
        
        # Expanded window to fit the new large form
        window_width = 750 
        window_height = 500
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        center_x = int(screen_width / 2 - window_width / 2)
        center_y = int(screen_height / 2 - window_height / 2)
        self.root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        
        self.setup_styles()
        self.setup_files()
        
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(pady=15, padx=15, fill='both', expand=True)
        
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        self.status_bar = tk.Label(self.root, textvariable=self.status_var, bg="#f4f5f7", fg="#5e6c84", font=("Segoe UI", 9), anchor="w")
        self.status_bar.pack(side="bottom", fill="x", padx=15, pady=5)

        self.create_widgets()

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        
        default_font = ("Segoe UI", 10)
        style.configure(".", font=default_font, background="#ffffff")
        style.configure("TNotebook", background="#f4f5f7", borderwidth=0)
        style.configure("TNotebook.Tab", background="#dfe1e6", foreground="#172b4d", padding=[15, 5], font=("Segoe UI", 10, "bold"))
        style.map("TNotebook.Tab", background=[("selected", "#ffffff")], foreground=[("selected", "#0052cc")])
        
        style.configure("TLabel", background="#ffffff", foreground="#172b4d", font=("Segoe UI", 10, "bold"))
        
        style.configure("Primary.TButton", background="#0052cc", foreground="white", font=("Segoe UI", 10, "bold"), padding=5)
        style.map("Primary.TButton", background=[("active", "#003d99")])
        
        style.configure("Danger.TButton", background="#de350b", foreground="white", font=("Segoe UI", 10, "bold"), padding=5)
        style.map("Danger.TButton", background=[("active", "#bf2600")])

        style.configure("Neutral.TButton", background="#ebecf0", foreground="#172b4d", font=("Segoe UI", 10, "bold"), padding=5)
        style.map("Neutral.TButton", background=[("active", "#dfe1e6")])

        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#fafbfc", foreground="#5e6c84")
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=25)

    def setup_files(self):
        """Creates the Excel and TXT files with expanded headers."""
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Accounts"
            # Expanded Headers (Balance is now Column 8)
            ws.append(["Account ID", "Full Name", "DOB", "Gender", "Phone", "Account Type", "Nominee", "Balance"])
            wb.save(EXCEL_FILE)

        if not os.path.exists(TXT_FILE):
            with open(TXT_FILE, "w", encoding="utf-8") as f:
                f.write("Date|Account ID|Action|Amount|Balance\n")

    def show_status(self, message, msg_type="info"):
        colors = {"info": "#0052cc", "success": "#00875a", "error": "#de350b"}
        self.status_bar.config(fg=colors.get(msg_type, "#5e6c84"))
        self.status_var.set(message)
        self.root.after(5000, lambda: [self.status_var.set("Ready"), self.status_bar.config(fg="#5e6c84")])

    def add_placeholder(self, entry, placeholder_text):
        entry.insert(0, placeholder_text)
        entry.config(foreground="#a5adba")

        def on_focus_in(event):
            if entry.get() == placeholder_text:
                entry.delete(0, tk.END)
                entry.config(foreground="#172b4d")

        def on_focus_out(event):
            if not entry.get():
                entry.insert(0, placeholder_text)
                entry.config(foreground="#a5adba")

        entry.bind("<FocusIn>", on_focus_in)
        entry.bind("<FocusOut>", on_focus_out)

    def log_transaction(self, acc_id, action, amount, balance):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        with open(TXT_FILE, "a", encoding="utf-8") as f:
            f.write(f"{timestamp}|{acc_id}|{action}|₹{amount:,.2f}|₹{balance:,.2f}\n")
        self.load_history()

    def create_widgets(self):
        self.frame_create = ttk.Frame(self.notebook)
        self.frame_transact = ttk.Frame(self.notebook)
        self.frame_history = ttk.Frame(self.notebook)

        self.notebook.add(self.frame_create, text='  Create Account  ')
        self.notebook.add(self.frame_transact, text='  Transactions  ')
        self.notebook.add(self.frame_history, text='  History  ')

        self._build_create_tab()
        self._build_transact_tab()
        self._build_history_tab()

    def _build_create_tab(self):
        # --- COLUMN 1 (Left Side) ---
        ttk.Label(self.frame_create, text="Full Name").place(x=40, y=20)
        self.entry_name = ttk.Entry(self.frame_create, width=35, font=("Segoe UI", 11))
        self.entry_name.place(x=40, y=50)
        self.add_placeholder(self.entry_name, "e.g., John Doe")

        ttk.Label(self.frame_create, text="Date of Birth").place(x=40, y=100)
        self.entry_dob = ttk.Entry(self.frame_create, width=35, font=("Segoe UI", 11))
        self.entry_dob.place(x=40, y=130)
        self.add_placeholder(self.entry_dob, "DD/MM/YYYY")

        ttk.Label(self.frame_create, text="Contact Number").place(x=40, y=180)
        self.entry_phone = ttk.Entry(self.frame_create, width=35, font=("Segoe UI", 11))
        self.entry_phone.place(x=40, y=210)
        self.add_placeholder(self.entry_phone, "10-digit mobile number")

        ttk.Label(self.frame_create, text="Initial Deposit (₹)").place(x=40, y=260)
        self.entry_initial = ttk.Entry(self.frame_create, width=35, font=("Segoe UI", 11))
        self.entry_initial.place(x=40, y=290)
        self.add_placeholder(self.entry_initial, "Minimum ₹1000")

        # --- COLUMN 2 (Right Side) ---
        ttk.Label(self.frame_create, text="Gender").place(x=380, y=20)
        self.combo_gender = ttk.Combobox(self.frame_create, values=["Male", "Female", "Other"], state="readonly", width=33, font=("Segoe UI", 11))
        self.combo_gender.place(x=380, y=50)
        self.combo_gender.set("Select Gender")

        ttk.Label(self.frame_create, text="Account Type").place(x=380, y=100)
        self.combo_type = ttk.Combobox(self.frame_create, values=["Savings", "Current", "Salary"], state="readonly", width=33, font=("Segoe UI", 11))
        self.combo_type.place(x=380, y=130)
        self.combo_type.set("Select Account Type")

        ttk.Label(self.frame_create, text="Nominee Name").place(x=380, y=180)
        self.entry_nominee = ttk.Entry(self.frame_create, width=35, font=("Segoe UI", 11))
        self.entry_nominee.place(x=380, y=210)
        self.add_placeholder(self.entry_nominee, "e.g., Jane Doe")

        # Submit Button
        ttk.Button(self.frame_create, text="Create New Account", style="Primary.TButton", cursor="hand2", command=self.create_account).place(x=380, y=280, width=290, height=40)

    def _build_transact_tab(self):
        ttk.Label(self.frame_transact, text="Account ID").place(x=40, y=40)
        self.entry_acc_id = ttk.Entry(self.frame_transact, width=40, font=("Segoe UI", 11))
        self.entry_acc_id.place(x=40, y=70)
        self.add_placeholder(self.entry_acc_id, "e.g., 101")

        ttk.Label(self.frame_transact, text="Amount (₹)").place(x=40, y=120)
        self.entry_amount = ttk.Entry(self.frame_transact, width=40, font=("Segoe UI", 11))
        self.entry_amount.place(x=40, y=150)
        self.add_placeholder(self.entry_amount, "e.g., 500")

        ttk.Button(self.frame_transact, text="Deposit", style="Primary.TButton", cursor="hand2", command=lambda: self.process_transaction("Deposit")).place(x=40, y=210, width=90)
        ttk.Button(self.frame_transact, text="Withdraw", style="Danger.TButton", cursor="hand2", command=lambda: self.process_transaction("Withdraw")).place(x=140, y=210, width=90)
        ttk.Button(self.frame_transact, text="Check Balance", style="Neutral.TButton", cursor="hand2", command=self.check_balance).place(x=240, y=210, width=120)
        ttk.Button(self.frame_transact, text="Delete Account", style="Danger.TButton", cursor="hand2", command=self.delete_account).place(x=400, y=210, width=130)

    def _build_history_tab(self):
        columns = ("Date", "ID", "Action", "Amount", "Balance")
        self.tree = ttk.Treeview(self.frame_history, columns=columns, show="headings", height=13)
        
        self.tree.heading("Date", text="Date & Time")
        self.tree.column("Date", width=140, anchor="center")
        self.tree.heading("ID", text="Acc ID")
        self.tree.column("ID", width=70, anchor="center")
        self.tree.heading("Action", text="Action")
        self.tree.column("Action", width=110, anchor="center")
        self.tree.heading("Amount", text="Amount")
        self.tree.column("Amount", width=100, anchor="e")
        self.tree.heading("Balance", text="Balance")
        self.tree.column("Balance", width=100, anchor="e")

        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        scrollbar = ttk.Scrollbar(self.tree, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        self.load_history()

    # --- Core Logic Methods ---

    def create_account(self):
        # Gather all data
        name = self.entry_name.get().strip()
        dob = self.entry_dob.get().strip()
        phone = self.entry_phone.get().strip()
        gender = self.combo_gender.get()
        acc_type = self.combo_type.get()
        nominee = self.entry_nominee.get().strip()
        init_val = self.entry_initial.get().strip()

        # Basic Validation to ensure no empty fields
        placeholders = ["e.g., John Doe", "DD/MM/YYYY", "10-digit mobile number", "e.g., Jane Doe", "Minimum ₹1000"]
        if name in placeholders or dob in placeholders or phone in placeholders or init_val in placeholders:
            self.show_status("Please fill in all text fields correctly.", "error")
            return
        
        if gender == "Select Gender" or acc_type == "Select Account Type":
            self.show_status("Please select a valid Gender and Account Type.", "error")
            return

        try:
            initial_deposit = float(init_val)
        except ValueError:
            self.show_status("Invalid deposit amount. Numbers only.", "error")
            return

        if initial_deposit < 0:
            self.show_status("Deposit cannot be negative.", "error")
            return

        # Write to Database
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        acc_id = 100 + ws.max_row
        # Notice how we append all 8 columns of data now
        ws.append([acc_id, name, dob, gender, phone, acc_type, nominee, initial_deposit])
        wb.save(EXCEL_FILE)
        
        self.log_transaction(acc_id, "Creation", initial_deposit, initial_deposit)
        self.show_status(f"Account created! ID: {acc_id} | Type: {acc_type}", "success")
        
        # Reset fields
        self.entry_name.delete(0, tk.END)
        self.entry_dob.delete(0, tk.END)
        self.entry_phone.delete(0, tk.END)
        self.entry_nominee.delete(0, tk.END)
        self.entry_initial.delete(0, tk.END)
        self.combo_gender.set("Select Gender")
        self.combo_type.set("Select Account Type")
        
        self.add_placeholder(self.entry_name, "e.g., John Doe")
        self.add_placeholder(self.entry_dob, "DD/MM/YYYY")
        self.add_placeholder(self.entry_phone, "10-digit mobile number")
        self.add_placeholder(self.entry_nominee, "e.g., Jane Doe")
        self.add_placeholder(self.entry_initial, "Minimum ₹1000")

    def process_transaction(self, action):
        id_val = self.entry_acc_id.get().strip()
        amt_val = self.entry_amount.get().strip()

        try:
            acc_id = int(id_val)
            amount = float(amt_val)
        except ValueError:
            self.show_status("Please enter valid numeric IDs and amounts.", "error")
            return

        if amount <= 0:
            self.show_status("Amount must be greater than zero.", "error")
            return

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        found = False

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == acc_id:
                found = True
                # Balance is now in Column 8
                current_balance = float(ws.cell(row=row, column=8).value)
                
                if action == "Withdraw" and current_balance < amount:
                    self.show_status(f"Insufficient funds. Balance: ₹{current_balance:,.2f}", "error")
                    return
                
                new_balance = current_balance + amount if action == "Deposit" else current_balance - amount
                ws.cell(row=row, column=8).value = new_balance
                wb.save(EXCEL_FILE)
                
                self.log_transaction(acc_id, action, amount, new_balance)
                self.show_status(f"{action} successful. New Balance: ₹{new_balance:,.2f}", "success")
                
                self.entry_amount.delete(0, tk.END)
                self.add_placeholder(self.entry_amount, "e.g., 500")
                break

        if not found:
            self.show_status("Account ID not found.", "error")

    def check_balance(self):
        try:
            acc_id = int(self.entry_acc_id.get().strip())
        except ValueError:
            self.show_status("Please enter a valid numeric Account ID.", "error")
            return

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == acc_id:
                name = ws.cell(row=row, column=2).value
                acc_type = ws.cell(row=row, column=6).value
                # Balance is now in Column 8
                balance = float(ws.cell(row=row, column=8).value)
                
                self.show_status(f"{name} ({acc_type}) Balance: ₹{balance:,.2f}", "info")
                return
                
        self.show_status("Account ID not found.", "error")

    def delete_account(self):
        try:
            acc_id = int(self.entry_acc_id.get().strip())
        except ValueError:
            self.show_status("Please enter a valid numeric Account ID.", "error")
            return

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        found_row = None

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == acc_id:
                found_row = row
                break

        if found_row:
            name = ws.cell(row=found_row, column=2).value
            # Balance is now in Column 8
            balance = float(ws.cell(row=found_row, column=8).value)
            
            confirm = messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete {name}'s account?\nThe remaining balance of ₹{balance:,.2f} will be closed.")
            
            if confirm:
                ws.delete_rows(found_row, 1)
                wb.save(EXCEL_FILE)
                
                self.log_transaction(acc_id, "Deleted", balance, 0.0)
                self.show_status(f"Account {acc_id} deleted successfully.", "success")
                
                self.entry_acc_id.delete(0, tk.END)
                self.add_placeholder(self.entry_acc_id, "e.g., 101")
            else:
                self.show_status("Account deletion cancelled.", "info")
        else:
            self.show_status("Account ID not found.", "error")

    def load_history(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        if not os.path.exists(TXT_FILE):
            return

        with open(TXT_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()[1:]
            for line in reversed(lines):
                data = line.strip().split("|")
                if len(data) == 5:
                    self.tree.insert("", "end", values=data)

if __name__ == "__main__":
    root = tk.Tk()
    app = BankApp(root)
    root.mainloop()